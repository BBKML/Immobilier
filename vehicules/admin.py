from django.contrib import admin
from django.contrib.admin import AdminSite
from django.db.models import Count, Q, Sum
from django.utils.html import format_html, mark_safe
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.http import JsonResponse
from django.db.models.functions import TruncMonth, TruncYear
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .models import Client, Proprietaire, Marque, TypeVehicule, TypeMoto, Vehicule, Moto
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

# Fonction utilitaire pour tronquer le texte
def truncate_text(text, max_length=15):
    """Tronque le texte si il dépasse la longueur maximale"""
    if text and len(str(text)) > max_length:
        return str(text)[:max_length-3] + '...'
    return str(text) if text else '-'

# Fonction utilitaire pour vérifier si les tables existent
def tables_exist():
    """Vérifier si les tables de base de données existent"""
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='vehicules_vehicule';")
            return cursor.fetchone() is not None
    except:
        return False

# Fonction utilitaire pour accéder aux modèles de manière sécurisée
def safe_count(model_class, **filters):
    """Compter les objets d'un modèle de manière sécurisée"""
    if not tables_exist():
        return 0
    try:
        return model_class.objects.filter(**filters).count()
    except:
        return 0

def safe_queryset(model_class, **filters):
    """Obtenir un queryset de manière sécurisée"""
    if not tables_exist():
        return model_class.objects.none()
    try:
        return model_class.objects.filter(**filters)
    except:
        return model_class.objects.none()

# Actions en lot personnalisées
@admin.action(description="Marquer comme actif")
def marquer_actif(modeladmin, request, queryset):
    queryset.update(statut='actif')

@admin.action(description="Marquer comme inactif")
def marquer_inactif(modeladmin, request, queryset):
    queryset.update(statut='inactif')

@admin.action(description="Exporter les données sélectionnées en CSV")
def exporter_donnees_csv(modeladmin, request, queryset):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{modeladmin.model._meta.verbose_name_plural}_selection.csv"'
    
    writer = csv.writer(response)
    
    # En-têtes selon le modèle
    if modeladmin.model == Vehicule:
        writer.writerow(['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut'])
        for obj in queryset:
            writer.writerow([
                obj.immatriculation,
                obj.marque.nom if obj.marque else '',
                obj.chassis,
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.client.nom if obj.client else '',
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Moto:
        writer.writerow(['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut'])
        for obj in queryset:
            writer.writerow([
                obj.immatriculation,
                obj.marque.nom if obj.marque else '',
                obj.chassis,
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.client.nom if obj.client else '',
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Client:
        writer.writerow(['Nom', 'Contact', 'Adresse'])
        for obj in queryset:
            writer.writerow([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Proprietaire:
        writer.writerow(['Nom', 'Contact', 'Adresse'])
        for obj in queryset:
            writer.writerow([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Marque:
        writer.writerow(['Nom', 'Catégorie'])
        for obj in queryset:
            writer.writerow([obj.nom, obj.categorie])
    elif modeladmin.model == TypeVehicule:
        writer.writerow(['Nom'])
        for obj in queryset:
            writer.writerow([obj.nom])
    elif modeladmin.model == TypeMoto:
        writer.writerow(['Nom'])
        for obj in queryset:
            writer.writerow([obj.nom])
    
    return response

@admin.action(description="Exporter les données sélectionnées en Excel")
def exporter_donnees_excel(modeladmin, request, queryset):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = modeladmin.model._meta.verbose_name_plural
    
    # En-têtes selon le modèle
    if modeladmin.model == Vehicule:
        headers = ['NOM DU PROPRIETAIRE', 'CONTACT PROPRIETAIRE', 'MARQUE', 'CHASSIS', 'CLIENTS', 'CONTACT CLIENT', 'NO IMM', 'DATES', 'CHRONO', 'TYPE TECH','STATUT.']
        ws.append(headers)
        for obj in queryset:
            ws.append([
                obj.proprietaire.nom if obj.proprietaire and obj.proprietaire.nom else '',
                obj.proprietaire.contact if obj.proprietaire and obj.proprietaire.contact else '',
                obj.marque.nom if obj.marque and obj.marque.nom else '',
                obj.chassis if obj.chassis else '',
                obj.client.nom if obj.client and obj.client.nom else '',
                obj.client.contact if obj.client and obj.client.contact else '',
                obj.immatriculation if obj.immatriculation else '',
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.chrono if obj.chrono else '',
                obj.type_tech if obj.type_tech else '',
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Moto:
        headers = ['NOM DU PROPRIETAIRE', 'CONTACT PROPRIETAIRE', 'MARQUE', 'CHASSIS', 'CLIENTS', 'CONTACT CLIENT', 'NO IMM', 'DATES', 'CHRONO', 'TYPE TECH','STATUT.']
        ws.append(headers)
        for obj in queryset:
            ws.append([
                obj.proprietaire.nom if obj.proprietaire and obj.proprietaire.nom else '',
                obj.proprietaire.contact if obj.proprietaire and obj.proprietaire.contact else '',
                obj.marque.nom if obj.marque and obj.marque.nom else '',
                obj.chassis if obj.chassis else '',
                obj.client.nom if obj.client and obj.client.nom else '',
                obj.client.contact if obj.client and obj.client.contact else '',
                obj.immatriculation if obj.immatriculation else '',
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.chrono if obj.chrono else '',
                obj.type_tech if obj.type_tech else '',
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Client:
        headers = ['Nom', 'Contact', 'Adresse']
        ws.append(headers)
        for obj in queryset:
            ws.append([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Proprietaire:
        headers = ['Nom', 'Contact', 'Adresse']
        ws.append(headers)
        for obj in queryset:
            ws.append([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Marque:
        headers = ['Nom', 'Catégorie']
        ws.append(headers)
        for obj in queryset:
            ws.append([obj.nom, obj.categorie])
    elif modeladmin.model == TypeVehicule:
        headers = ['Nom']
        ws.append(headers)
        for obj in queryset:
            ws.append([obj.nom])
    elif modeladmin.model == TypeMoto:
        headers = ['Nom']
        ws.append(headers)
        for obj in queryset:
            ws.append([obj.nom])
    
    # Style en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{modeladmin.model._meta.verbose_name_plural}_selection.xlsx"'
    return response

@admin.action(description="Exporter les données sélectionnées en PDF")
def exporter_donnees_pdf(modeladmin, request, queryset):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from io import BytesIO
    from django.http import HttpResponse

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                          leftMargin=0.1*inch, rightMargin=0.1*inch, 
                          topMargin=0.2*inch, bottomMargin=0.2*inch)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=15,
        alignment=1
    )
    title = Paragraph(f"Liste des {modeladmin.model._meta.verbose_name_plural} sélectionnés", title_style)
    elements.append(title)
    elements.append(Spacer(1, 15))

    # En-têtes selon le modèle
    if modeladmin.model == Vehicule:
        headers = ['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.immatriculation,
                truncate_text(obj.marque.nom if obj.marque else '-', 10),
                truncate_text(obj.chassis, 15),
                str(obj.date)[:10] if obj.date else '-',
                truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15),
                truncate_text(obj.client.nom if obj.client else '-', 15),
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Moto:
        headers = ['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.immatriculation,
                truncate_text(obj.marque.nom if obj.marque else '-', 10),
                truncate_text(obj.chassis, 15),
                str(obj.date)[:10] if obj.date else '-',
                truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15),
                truncate_text(obj.client.nom if obj.client else '-', 15),
                getattr(obj, 'statut', 'actif')
            ])
    elif modeladmin.model == Client:
        headers = ['Nom', 'Contact', 'Adresse']
        data = [headers]
        for obj in queryset:
            data.append([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Proprietaire:
        headers = ['Nom', 'Contact', 'Adresse']
        data = [headers]
        for obj in queryset:
            data.append([obj.nom, obj.contact, obj.adresse])
    elif modeladmin.model == Marque:
        headers = ['Nom', 'Catégorie']
        data = [headers]
        for obj in queryset:
            data.append([obj.nom, obj.categorie])
    elif modeladmin.model == TypeVehicule:
        headers = ['Nom']
        data = [headers]
        for obj in queryset:
            data.append([obj.nom])
    elif modeladmin.model == TypeMoto:
        headers = ['Nom']
        data = [headers]
        for obj in queryset:
            data.append([obj.nom])
    
    # Calculer automatiquement les largeurs de colonnes
    available_width = landscape(A4)[0] - 0.2*inch
    num_columns = len(headers)
    base_width = available_width / num_columns
    
    col_widths = [base_width] * num_columns
    
    table = Table(data, colWidths=col_widths)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(table_style)
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{modeladmin.model._meta.verbose_name_plural}_selection.pdf"'
    response.write(pdf)
    return response

# Filtres personnalisés
class StatutFilter(admin.SimpleListFilter):
    title = 'Statut'
    parameter_name = 'statut'

    def lookups(self, request, model_admin):
        return (
            ('actif', 'Actif'),
            ('inactif', 'Inactif'),
            ('en_maintenance', 'En maintenance'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(statut=self.value())

class DateRangeFilter(admin.SimpleListFilter):
    title = 'Période'
    parameter_name = 'periode'

    def lookups(self, request, model_admin):
        return (
            ('aujourd_hui', 'Aujourd\'hui'),
            ('cette_semaine', 'Cette semaine'),
            ('ce_mois', 'Ce mois'),
            ('cette_annee', 'Cette année'),
        )

    def queryset(self, request, queryset):
        today = timezone.now().date()
        if self.value() == 'aujourd_hui':
            return queryset.filter(date=today)
        elif self.value() == 'cette_semaine':
            start_of_week = today - timedelta(days=today.weekday())
            return queryset.filter(date__gte=start_of_week)
        elif self.value() == 'ce_mois':
            return queryset.filter(date__month=today.month, date__year=today.year)
        elif self.value() == 'cette_annee':
            return queryset.filter(date__year=today.year)

# Modernisation des formulaires admin pour tous les modèles

@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = ('nom', 'contact', 'adresse', 'vehicules_count', 'derniere_activite', 'actions_rapides')
    list_filter = (DateRangeFilter,)
    search_fields = ('nom', 'contact', 'adresse')
    list_per_page = 10
    ordering = ('nom',)
    actions = [exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Informations personnelles", {
            'fields': ('nom', 'contact', 'adresse'),
            'description': "Coordonnées du client.",
            'classes': ('wide',),
        }),
    )
    def vehicules_count(self, obj):
        return safe_count(Vehicule, client=obj) + safe_count(Moto, client=obj)
    vehicules_count.short_description = 'Véhicules'
    def derniere_activite(self, obj):
        return "-"
    derniere_activite.short_description = 'Dernière activité'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer ce client ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_client_change', args=[obj.pk]),
            reverse('admin:vehicules_client_change', args=[obj.pk]),
            reverse('admin:vehicules_client_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['nom'].help_text = "Nom complet du client."
        form.base_fields['contact'].help_text = "Numéro de téléphone ou email."
        form.base_fields['adresse'].help_text = "Adresse postale ou géographique."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV pour les clients"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="clients_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Contact', 'Adresse', 'Nombre de véhicules'])
        
        for obj in queryset:
            writer.writerow([
                obj.nom,
                obj.contact,
                obj.adresse,
                safe_count(Vehicule, client=obj) + safe_count(Moto, client=obj)
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel pour les clients"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        headers = ['NOM', 'CONTACT', 'ADRESSE', 'NOMBRE DE VÉHICULES']
        ws.append(headers)
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.nom if obj.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.contact if obj.contact else '')
            ws.cell(row=row_idx, column=3, value=obj.adresse if obj.adresse else '')
            ws.cell(row=row_idx, column=4, value=safe_count(Vehicule, client=obj) + safe_count(Moto, client=obj))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="clients_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF pour les clients"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch, 
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1
        )
        title = Paragraph("Liste des Clients", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        headers = ['Nom', 'Contact', 'Adresse', 'Véhicules']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.nom,
                obj.contact,
                truncate_text(obj.adresse, 30),
                str(safe_count(Vehicule, client=obj) + safe_count(Moto, client=obj))
            ])
        
        table = Table(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="clients_export.pdf"'
        return response

@admin.register(Proprietaire)
class ProprietaireAdmin(admin.ModelAdmin):
    list_display = ('nom', 'contact', 'adresse', 'vehicules_count', 'valeur_totale', 'actions_rapides')
    list_filter = (DateRangeFilter,)
    search_fields = ('nom', 'contact', 'adresse')
    list_per_page = 10
    ordering = ('nom',)
    actions = [exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Informations personnelles", {
            'fields': ('nom', 'contact', 'adresse'),
            'description': "Coordonnées du propriétaire.",
            'classes': ('wide',),
        }),
    )
    def vehicules_count(self, obj):
        return safe_count(Vehicule, proprietaire=obj) + safe_count(Moto, proprietaire=obj)
    vehicules_count.short_description = 'Véhicules'
    def valeur_totale(self, obj):
        return "-"
    valeur_totale.short_description = 'Valeur totale'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer ce propriétaire ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_proprietaire_change', args=[obj.pk]),
            reverse('admin:vehicules_proprietaire_change', args=[obj.pk]),
            reverse('admin:vehicules_proprietaire_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['nom'].help_text = "Nom complet du propriétaire."
        form.base_fields['contact'].help_text = "Numéro de téléphone ou email."
        form.base_fields['adresse'].help_text = "Adresse postale ou géographique."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV pour les propriétaires"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="proprietaires_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Contact', 'Adresse', 'Nombre de véhicules'])
        
        for obj in queryset:
            writer.writerow([
                obj.nom,
                obj.contact,
                obj.adresse,
                safe_count(Vehicule, proprietaire=obj) + safe_count(Moto, proprietaire=obj)
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel pour les propriétaires"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Propriétaires"
        
        headers = ['NOM', 'CONTACT', 'ADRESSE', 'NOMBRE DE VÉHICULES']
        ws.append(headers)
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.nom if obj.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.contact if obj.contact else '')
            ws.cell(row=row_idx, column=3, value=obj.adresse if obj.adresse else '')
            ws.cell(row=row_idx, column=4, value=safe_count(Vehicule, proprietaire=obj) + safe_count(Moto, proprietaire=obj))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="proprietaires_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF pour les propriétaires"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch, 
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1
        )
        title = Paragraph("Liste des Propriétaires", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        headers = ['Nom', 'Contact', 'Adresse', 'Véhicules']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.nom,
                obj.contact,
                truncate_text(obj.adresse, 30),
                str(safe_count(Vehicule, proprietaire=obj) + safe_count(Moto, proprietaire=obj))
            ])
        
        table = Table(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="proprietaires_export.pdf"'
        return response

@admin.register(Marque)
class MarqueAdmin(admin.ModelAdmin):
    list_display = ('nom', 'categorie', 'vehicules_count', 'motos_count', 'popularite', 'actions_rapides')
    list_filter = ('categorie',)
    search_fields = ('nom',)
    list_per_page = 10
    ordering = ('nom',)
    actions = [exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Informations de base", {
            'fields': ('nom', 'categorie'),
            'description': "Nom et catégorie de la marque.",
            'classes': ('wide',),
        }),
    )
    def vehicules_count(self, obj):
        return safe_count(Vehicule, marque=obj)
    vehicules_count.short_description = 'Véhicules'
    def motos_count(self, obj):
        return safe_count(Moto, marque=obj)
    motos_count.short_description = 'Motos'
    def popularite(self, obj):
        total = safe_count(Vehicule, marque=obj) + safe_count(Moto, marque=obj)
        if total > 10:
            return format_html('<span class="badge badge-success">Populaire</span>')
        elif total > 5:
            return format_html('<span class="badge badge-warning">Moyen</span>')
        else:
            return format_html('<span class="badge badge-secondary">Rare</span>')
    popularite.short_description = 'Popularité'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer cette marque ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_marque_change', args=[obj.pk]),
            reverse('admin:vehicules_marque_change', args=[obj.pk]),
            reverse('admin:vehicules_marque_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['nom'].help_text = "Nom de la marque."
        form.base_fields['categorie'].help_text = "Catégorie de la marque (auto, moto, etc.)."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV pour les marques"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="marques_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Catégorie', 'Véhicules', 'Motos', 'Total'])
        
        for obj in queryset:
            vehicules = safe_count(Vehicule, marque=obj)
            motos = safe_count(Moto, marque=obj)
            writer.writerow([
                obj.nom,
                obj.categorie,
                vehicules,
                motos,
                vehicules + motos
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel pour les marques"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Marques"
        
        headers = ['NOM', 'CATÉGORIE', 'VÉHICULES', 'MOTOS', 'TOTAL']
        ws.append(headers)
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, obj in enumerate(queryset, 2):
            vehicules = safe_count(Vehicule, marque=obj)
            motos = safe_count(Moto, marque=obj)
            total = vehicules + motos
            
            ws.cell(row=row_idx, column=1, value=obj.nom if obj.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.categorie if obj.categorie else '')
            ws.cell(row=row_idx, column=3, value=vehicules)
            ws.cell(row=row_idx, column=4, value=motos)
            ws.cell(row=row_idx, column=5, value=total)
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="marques_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF pour les marques"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch, 
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1
        )
        title = Paragraph("Liste des Marques", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        headers = ['Nom', 'Catégorie', 'Véhicules', 'Motos', 'Total']
        data = [headers]
        for obj in queryset:
            vehicules = safe_count(Vehicule, marque=obj)
            motos = safe_count(Moto, marque=obj)
            data.append([
                obj.nom,
                obj.categorie,
                str(vehicules),
                str(motos),
                str(vehicules + motos)
            ])
        
        table = Table(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="marques_export.pdf"'
        return response

@admin.register(TypeVehicule)
class TypeVehiculeAdmin(admin.ModelAdmin):
    list_display = ('nom', 'vehicules_count', 'actions_rapides')
    search_fields = ('nom',)
    list_per_page = 10
    ordering = ('nom',)
    actions = [exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Type de véhicule", {
            'fields': ('nom',),
            'description': "Nom du type de véhicule.",
            'classes': ('wide',),
        }),
    )
    def vehicules_count(self, obj):
        return safe_count(Vehicule, type_vehicule=obj)
    vehicules_count.short_description = 'Véhicules'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer ce type de véhicule ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_typevehicule_change', args=[obj.pk]),
            reverse('admin:vehicules_typevehicule_change', args=[obj.pk]),
            reverse('admin:vehicules_typevehicule_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['nom'].help_text = "Nom du type de véhicule (ex: Camion, Voiture, etc.)."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV pour les types de véhicules"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="types_vehicules_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Nombre de véhicules'])
        
        for obj in queryset:
            writer.writerow([
                obj.nom,
                safe_count(Vehicule, type_vehicule=obj)
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel pour les types de véhicules"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Types de Véhicules"
        
        headers = ['NOM', 'NOMBRE DE VÉHICULES']
        ws.append(headers)
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.nom if obj.nom else '')
            ws.cell(row=row_idx, column=2, value=safe_count(Vehicule, type_vehicule=obj))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="types_vehicules_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF pour les types de véhicules"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch, 
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1
        )
        title = Paragraph("Liste des Types de Véhicules", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        headers = ['Nom', 'Véhicules']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.nom,
                str(safe_count(Vehicule, type_vehicule=obj))
            ])
        
        table = Table(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="types_vehicules_export.pdf"'
        return response

@admin.register(TypeMoto)
class TypeMotoAdmin(admin.ModelAdmin):
    list_display = ('nom', 'motos_count', 'actions_rapides')
    search_fields = ('nom',)
    list_per_page = 10
    ordering = ('nom',)
    actions = [exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Type de moto", {
            'fields': ('nom',),
            'description': "Nom du type de moto.",
            'classes': ('wide',),
        }),
    )
    def motos_count(self, obj):
        return safe_count(Moto, type_moto=obj)
    motos_count.short_description = 'Motos'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer ce type de moto ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_typemoto_change', args=[obj.pk]),
            reverse('admin:vehicules_typemoto_change', args=[obj.pk]),
            reverse('admin:vehicules_typemoto_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['nom'].help_text = "Nom du type de moto (ex: Scooter, Cross, etc.)."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV pour les types de motos"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="types_motos_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Nombre de motos'])
        
        for obj in queryset:
            writer.writerow([
                obj.nom,
                safe_count(Moto, type_moto=obj)
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel pour les types de motos"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Types de Motos"
        
        headers = ['NOM', 'NOMBRE DE MOTOS']
        ws.append(headers)
        
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.nom if obj.nom else '')
            ws.cell(row=row_idx, column=2, value=safe_count(Moto, type_moto=obj))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="types_motos_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF pour les types de motos"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch, 
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1
        )
        title = Paragraph("Liste des Types de Motos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        headers = ['Nom', 'Motos']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.nom,
                str(safe_count(Moto, type_moto=obj))
            ])
        
        table = Table(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="types_motos_export.pdf"'
        return response

@admin.register(Vehicule)
class VehiculeAdmin(admin.ModelAdmin):
    list_display = ('immatriculation', 'marque_nom', 'chassis', 'date', 'proprietaire_nom', 'client_nom', 'statut_badge', 'actions_rapides')
    list_filter = ('marque', 'type_vehicule', 'date', 'client', 'proprietaire', StatutFilter, DateRangeFilter)
    search_fields = ('immatriculation', 'chassis', 'proprietaire__nom', 'client__nom', 'marque__nom')
    list_per_page = 25
    ordering = ('-date',)
    actions = [marquer_actif, marquer_inactif, exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Informations du véhicule", {
            'fields': ('marque', 'type_vehicule', 'chassis', 'immatriculation'),
            'description': "Informations du véhicule (obligatoires).",
            'classes': ('wide', 'extrapretty'),
        }),
        ("Informations administratives", {
            'fields': ('proprietaire', 'client', 'date', 'chrono', 'type_tech', 'statut'),
            'description': "Informations administratives (optionnelles - à compléter plus tard).",
            'classes': ('wide',),
        }),
    )
    def marque_nom(self, obj):
        return obj.marque.nom if obj.marque else '-'
    marque_nom.short_description = 'Marque'
    marque_nom.admin_order_field = 'marque__nom'
    def proprietaire_nom(self, obj):
        return obj.proprietaire.nom if obj.proprietaire else '-'
    proprietaire_nom.short_description = 'Propriétaire'
    proprietaire_nom.admin_order_field = 'proprietaire__nom'
    def client_nom(self, obj):
        return obj.client.nom if obj.client else '-'
    client_nom.short_description = 'Client'
    client_nom.admin_order_field = 'client__nom'
    def statut_badge(self, obj):
        if hasattr(obj, 'statut'):
            if obj.statut == 'actif':
                return format_html('<span class="badge badge-success">Actif</span>')
            elif obj.statut == 'inactif':
                return format_html('<span class="badge badge-secondary">Inactif</span>')
            else:
                return format_html('<span class="badge badge-secondary">{}</span>', obj.statut)
        return format_html('<span class="badge badge-success">Actif</span>')
    statut_badge.short_description = 'Statut'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer ce véhicule ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_vehicule_change', args=[obj.pk]),
            reverse('admin:vehicules_vehicule_change', args=[obj.pk]),
            reverse('admin:vehicules_vehicule_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['immatriculation'].help_text = "Numéro d'immatriculation du véhicule (ex: AA-123-BC)"
        form.base_fields['chassis'].help_text = "Numéro de châssis unique du véhicule."
        form.base_fields['marque'].help_text = "Sélectionnez la marque du véhicule."
        form.base_fields['type_vehicule'].help_text = "Type/catégorie du véhicule."
        form.base_fields['proprietaire'].help_text = "Propriétaire actuel du véhicule."
        form.base_fields['client'].help_text = "Client associé à ce véhicule."
        form.base_fields['date'].help_text = "Date d'enregistrement ou d'achat."
        form.base_fields['chrono'].help_text = "Numéro chrono administratif."
        form.base_fields['type_tech'].help_text = "Type technique ou référence technique."
        form.base_fields['statut'].help_text = "Statut actuel du véhicule (Actif, Inactif, En maintenance, etc.)."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_all_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_all_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_all_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="vehicules_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut'])
        
        for obj in queryset:
            writer.writerow([
                obj.immatriculation,
                obj.marque.nom if obj.marque else '',
                obj.chassis,
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.client.nom if obj.client else '',
                getattr(obj, 'statut', 'actif')
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel complet au format personnalisé et coloré pour les véhicules"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Véhicules"
        # En-têtes personnalisés
        headers = ['NOM DU PROPRIETAIRE', 'CONTACT PROPRIETAIRE', 'MARQUE', 'CHASSIS', 'CLIENTS', 'CONTACT CLIENT', 'NO IMM', 'DATES', 'CHRONO', 'TYPE TECH','STATUT.']
        ws.append(headers)
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Données
        chrono_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Vert clair
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.proprietaire.nom if obj.proprietaire and obj.proprietaire.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.proprietaire.contact if obj.proprietaire and obj.proprietaire.contact else '')
            ws.cell(row=row_idx, column=3, value=obj.marque.nom if obj.marque and obj.marque.nom else '')
            ws.cell(row=row_idx, column=4, value=obj.chassis if obj.chassis else '')
            ws.cell(row=row_idx, column=5, value=obj.client.nom if obj.client and obj.client.nom else '')
            ws.cell(row=row_idx, column=6, value=obj.client.contact if obj.client and obj.client.contact else '')
            ws.cell(row=row_idx, column=7, value=obj.immatriculation if obj.immatriculation else '')
            ws.cell(row=row_idx, column=8, value=obj.date.strftime('%d/%m/%Y') if obj.date else '')
            ws.cell(row=row_idx, column=9, value=obj.chrono if obj.chrono else '')
            ws.cell(row=row_idx, column=10, value=obj.type_tech if obj.type_tech else '')
            ws.cell(row=row_idx, column=11, value=getattr(obj, 'statut', 'actif'))

            # Appliquer le fond vert à la colonne CHRONO
            ws.cell(row=row_idx, column=9).fill = chrono_fill
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="vehicules_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF avec ReportLab (troncature et colonnes larges)"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        # Marges minimales pour maximiser l'espace d'affichage
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.1*inch, rightMargin=0.1*inch, 
                              topMargin=0.2*inch, bottomMargin=0.2*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=15,
            alignment=1
        )
        title = Paragraph("Liste des Véhicules", title_style)
        elements.append(title)
        elements.append(Spacer(1, 15))

        headers = ['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.immatriculation,
                truncate_text(obj.marque.nom if obj.marque else '-', 10),
                truncate_text(obj.chassis, 15),
                str(obj.date)[:10] if obj.date else '-',
                truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15),
                truncate_text(obj.client.nom if obj.client else '-', 15),
                getattr(obj, 'statut', 'actif')
            ])
        
        # Calculer automatiquement les largeurs de colonnes
        available_width = landscape(A4)[0] - 0.2*inch  # Largeur disponible moins les marges
        num_columns = len(headers)
        base_width = available_width / num_columns
        
        # Ajuster certaines colonnes selon leur contenu
        col_widths = []
        for i, header in enumerate(headers):
            if 'Immatriculation' in header:
                col_widths.append(base_width * 1.3)  # Plus large pour l'immatriculation
            elif 'Chassis' in header:
                col_widths.append(base_width * 1.4)  # Plus large pour le chassis
            elif 'Date' in header:
                col_widths.append(base_width * 0.8)  # Plus étroite pour la date
            elif 'Statut' in header:
                col_widths.append(base_width * 0.7)  # Plus étroite pour le statut
            else:
                col_widths.append(base_width)
        
        table = Table(data, colWidths=col_widths)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(table_style)
        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="vehicules_export.pdf"'
        response.write(pdf)
        return response
    
    def export_all_csv(self, request, queryset):
        """Export CSV complet au format personnalisé pour les véhicules"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="vehicules_complet_export.csv"'
        
        writer = csv.writer(response)
        # En-têtes personnalisés
        writer.writerow([
            'NOM DU PROPRIETAIRE', 'MARQUE', 'CHASSIS', 'CLIENTS', 'NO IMM', 'DATES', 'CHRONO', 'TYPE TECH.'
        ])
        for obj in queryset:
            writer.writerow([
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.marque.nom if obj.marque else '',
                obj.chassis,
                obj.client.nom if obj.client else '',
                obj.immatriculation,
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.chrono,
                obj.type_tech
            ])
        return response

    def export_all_pdf(self, request, queryset):
        """Export PDF complet avec toutes les données détaillées"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        # Marges minimales pour maximiser l'espace d'affichage
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.05*inch, rightMargin=0.05*inch, 
                              topMargin=0.1*inch, bottomMargin=0.1*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=12,
            spaceAfter=10,
            alignment=1
        )
        title = Paragraph("Export Complet - Véhicules", title_style)
        elements.append(title)
        elements.append(Spacer(1, 10))

        # En-têtes détaillés SANS 'Adresse Propriétaire' et 'Adresse Client'
        headers = [
            'Immatriculation', 'Marque', 'Type Véhicule', 'Chassis', 'Date', 
            'Propriétaire', 'Contact Propriétaire',
            'Client', 'Contact Client', 'Chrono', 'Type Tech'
        ]
        cell_style = ParagraphStyle('cell', fontSize=6, leading=7, alignment=1)
        data = [headers]
        for obj in queryset:
            data.append([
                obj.immatriculation,
                truncate_text(obj.marque.nom if obj.marque else '-', 10),
                truncate_text(obj.type_vehicule.nom if obj.type_vehicule else '-', 12),
                truncate_text(obj.chassis, 15),
                str(obj.date)[:10] if obj.date else '-',
                truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15),
                truncate_text(obj.proprietaire.contact if obj.proprietaire else '-', 15),
                truncate_text(obj.client.nom if obj.client else '-', 15),
                truncate_text(obj.client.contact if obj.client else '-', 15),
                truncate_text(obj.chrono, 10),
                obj.type_tech
            ])
        
        # Calculer automatiquement les largeurs de colonnes pour s'adapter à la page
        available_width = landscape(A4)[0] - 0.1*inch  # Largeur disponible moins les marges
        num_columns = len(headers)
        base_width = available_width / num_columns
        
        # Ajuster certaines colonnes selon leur contenu
        col_widths = []
        for i, header in enumerate(headers):
            if 'Immatriculation' in header:
                col_widths.append(base_width * 1.5)  # Plus large pour l'immatriculation
            elif 'Chassis' in header:
                col_widths.append(base_width * 1.3)  # Plus large pour le chassis
            elif 'Date' in header:
                col_widths.append(base_width * 0.8)  # Plus étroite pour la date
            elif 'Statut' in header:
                col_widths.append(base_width * 0.7)  # Plus étroite pour le statut
            else:
                col_widths.append(base_width)
        
        table = Table(data, colWidths=col_widths)
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(table_style)
        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="vehicules_complet_export.pdf"'
        response.write(pdf)
        return response

    def export_all_excel(self, request, queryset):
        """Export Complet Excel au format personnalisé et coloré pour les véhicules (même logique que export_excel)"""
        return self.export_excel(request, queryset)

@admin.register(Moto)
class MotoAdmin(admin.ModelAdmin):
    list_display = ('immatriculation', 'marque_nom', 'chassis', 'date', 'proprietaire_nom', 'client_nom', 'statut_badge', 'actions_rapides')
    list_filter = ('marque', 'type_moto', 'date', 'client', 'proprietaire', StatutFilter, DateRangeFilter)
    search_fields = ('immatriculation', 'chassis', 'proprietaire__nom', 'client__nom', 'marque__nom')
    list_per_page = 25
    ordering = ('-date',)
    actions = [marquer_actif, marquer_inactif, exporter_donnees_csv, exporter_donnees_excel, exporter_donnees_pdf]
    change_list_template = 'admin/vehicules/change_list.html'
    fieldsets = (
        ("Moto", {
            'fields': ('marque', 'type_moto', 'chassis', 'immatriculation', 'couleurs'),
            'description': "Informations de la moto (obligatoires).",
            'classes': ('wide', 'extrapretty'),
        }),
        ("Administration", {
            'fields': ('proprietaire', 'client', 'date', 'no_fiche', 'couleur', 'cellulaire', 'statut'),
            'description': "Informations administratives (optionnelles - à compléter plus tard).",
            'classes': ('wide',),
        }),
    )
    def marque_nom(self, obj):
        return obj.marque.nom if obj.marque else '-'
    marque_nom.short_description = 'Marque'
    marque_nom.admin_order_field = 'marque__nom'
    def proprietaire_nom(self, obj):
        return obj.proprietaire.nom if obj.proprietaire else '-'
    proprietaire_nom.short_description = 'Propriétaire'
    proprietaire_nom.admin_order_field = 'proprietaire__nom'
    def client_nom(self, obj):
        return obj.client.nom if obj.client else '-'
    client_nom.short_description = 'Client'
    client_nom.admin_order_field = 'client__nom'
    def statut_badge(self, obj):
        if hasattr(obj, 'statut'):
            if obj.statut == 'actif':
                return format_html('<span class="badge badge-success">Actif</span>')
            elif obj.statut == 'inactif':
                return format_html('<span class="badge badge-secondary">Inactif</span>')
            else:
                return format_html('<span class="badge badge-secondary">{}</span>', obj.statut)
        return format_html('<span class="badge badge-success">Actif</span>')
    statut_badge.short_description = 'Statut'
    def actions_rapides(self, obj):
        return format_html(
            '<div style="display: flex; gap: 5px; justify-content: center; align-items: center;">'
            '<a href="{}" class="btn btn-sm btn-primary" title="Voir" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-eye" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-warning" title="Modifier" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;">'
            '<i class="fas fa-edit" style="font-size: 12px;"></i></a>'
            '<a href="{}" class="btn btn-sm btn-danger" title="Supprimer" style="display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; padding: 0; border-radius: 4px;" onclick="return confirm(\'Êtes-vous sûr de vouloir supprimer cette moto ?\');">'
            '<i class="fas fa-trash" style="font-size: 12px;"></i></a>'
            '</div>',
            reverse('admin:vehicules_moto_change', args=[obj.pk]),
            reverse('admin:vehicules_moto_change', args=[obj.pk]),
            reverse('admin:vehicules_moto_delete', args=[obj.pk])
        )
    actions_rapides.short_description = 'Actions'
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['immatriculation'].help_text = "Numéro d'immatriculation de la moto."
        form.base_fields['chassis'].help_text = "Numéro de châssis unique."
        form.base_fields['marque'].help_text = "Sélectionnez la marque."
        form.base_fields['type_moto'].help_text = "Type/catégorie de la moto."
        form.base_fields['proprietaire'].help_text = "Propriétaire actuel."
        form.base_fields['client'].help_text = "Client associé."
        form.base_fields['date'].help_text = "Date d'enregistrement ou d'achat."
        form.base_fields['no_fiche'].help_text = "Numéro de fiche administrative."
        form.base_fields['couleur'].help_text = "Couleur de la moto."
        form.base_fields['couleurs'].help_text = "Autres couleurs de la moto."
        form.base_fields['cellulaire'].help_text = "Numéro de téléphone ou info cellulaire."
        form.base_fields['statut'].help_text = "Statut actuel de la moto (Actif, Inactif, En maintenance, etc.)."
        return form
    
    def changelist_view(self, request, extra_context=None):
        # Gérer les exports depuis les boutons
        export_type = request.GET.get('export')
        if export_type:
            # Vérifier s'il y a des éléments sélectionnés
            selected_ids = request.GET.getlist('_selected_action')
            if selected_ids:
                queryset = self.model.objects.filter(id__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_data(request, queryset, export_type)
        return super().changelist_view(request, extra_context)
    
    def export_data(self, request, queryset, export_type):
        """Gérer l'export des données selon le type demandé"""
        try:
            if export_type == 'csv':
                return self.export_csv(request, queryset)
            elif export_type == 'excel':
                return self.export_excel(request, queryset)
            elif export_type == 'pdf':
                return self.export_pdf(request, queryset)
            elif export_type == 'all_csv':
                return self.export_all_csv(request, queryset)
            elif export_type == 'all_pdf':
                return self.export_all_pdf(request, queryset)
            elif export_type == 'all_excel':
                return self.export_all_excel(request, queryset)
            else:
                from django.contrib import messages
                from django.shortcuts import redirect
                messages.error(request, f"Type d'export '{export_type}' non reconnu.")
                return redirect(request.path)
        except Exception as e:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Erreur lors de l'export : {str(e)}")
            return redirect(request.path)
    
    def export_csv(self, request, queryset):
        """Export CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="motos_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut'])
        
        for obj in queryset:
            writer.writerow([
                obj.immatriculation,
                obj.marque.nom if obj.marque else '',
                obj.chassis,
                obj.date.strftime('%d/%m/%Y') if obj.date else '',
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.client.nom if obj.client else '',
                getattr(obj, 'statut', 'actif')
            ])
        
        return response
    
    def export_excel(self, request, queryset):
        """Export Excel complet au format personnalisé et coloré pour les motos"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Motos"
        # En-têtes personnalisés
        headers = ['NOM DU PROPRIÉTAIRE', 'CONTACT PROPRIÉTAIRE', 'MARQUE', 'TYPE', 'N° CHASSIS', 'N° IMMAT.', 'DATES', 'NO FICHE', 'CLIENT', 'CONTACT CLIENT', 'COULEURS']
        ws.append(headers)
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Données
        immat_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Jaune clair
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.proprietaire.nom if obj.proprietaire and obj.proprietaire.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.proprietaire.contact if obj.proprietaire and obj.proprietaire.contact else '')
            ws.cell(row=row_idx, column=3, value=obj.marque.nom if obj.marque and obj.marque.nom else '')
            ws.cell(row=row_idx, column=4, value=obj.type_moto.nom if obj.type_moto and obj.type_moto.nom else '')
            ws.cell(row=row_idx, column=5, value=obj.chassis if obj.chassis else '')
            ws.cell(row=row_idx, column=6, value=obj.immatriculation if obj.immatriculation else '')
            ws.cell(row=row_idx, column=7, value=obj.date.strftime('%d/%m/%Y') if obj.date else '')
            ws.cell(row=row_idx, column=8, value=obj.no_fiche if obj.no_fiche else '')
            ws.cell(row=row_idx, column=9, value=obj.client.nom if obj.client and obj.client.nom else '')
            ws.cell(row=row_idx, column=10, value=obj.client.contact if obj.client and obj.client.contact else '')
            ws.cell(row=row_idx, column=11, value=obj.couleurs if obj.couleurs else '')
            

            # Appliquer le fond jaune à la colonne N° IMMAT.
            ws.cell(row=row_idx, column=6).fill = immat_fill
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="motos_complet_export.xlsx"'
        return response
    
    def export_pdf(self, request, queryset):
        """Export PDF simple pour les motos"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        # Marges minimales pour maximiser l'espace d'affichage
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.1*inch, rightMargin=0.1*inch, 
                              topMargin=0.2*inch, bottomMargin=0.2*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=15,
            alignment=1
        )
        title = Paragraph("Liste des Motos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 15))

        headers = ['Immatriculation', 'Marque', 'Chassis', 'Date', 'Propriétaire', 'Client', 'Statut']
        data = [headers]
        for obj in queryset:
            data.append([
                obj.immatriculation,
                truncate_text(obj.marque.nom if obj.marque else '-', 10),
                truncate_text(obj.chassis, 15),
                str(obj.date)[:10] if obj.date else '-',
                truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15),
                truncate_text(obj.client.nom if obj.client else '-', 15),
                getattr(obj, 'statut', 'actif')
            ])
        
        # Calculer automatiquement les largeurs de colonnes
        available_width = landscape(A4)[0] - 0.2*inch  # Largeur disponible moins les marges
        num_columns = len(headers)
        base_width = available_width / num_columns
        
        # Ajuster certaines colonnes selon leur contenu
        col_widths = []
        for i, header in enumerate(headers):
            if 'Immatriculation' in header:
                col_widths.append(base_width * 1.3)  # Plus large pour l'immatriculation
            elif 'Chassis' in header:
                col_widths.append(base_width * 1.4)  # Plus large pour le chassis
            elif 'Date' in header:
                col_widths.append(base_width * 0.8)  # Plus étroite pour la date
            elif 'Statut' in header:
                col_widths.append(base_width * 0.7)  # Plus étroite pour le statut
            else:
                col_widths.append(base_width)
        
        table = Table(data, colWidths=col_widths)
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="motos_export.pdf"'
        response.write(pdf)
        
        return response
    
    def export_all_csv(self, request, queryset):
        """Export CSV complet avec toutes les données pour les motos"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="motos_complet_export.csv"'
        
        writer = csv.writer(response)
        
        # En-têtes détaillés SANS Statut et SANS adresses
        writer.writerow([
            'Immatriculation', 'Marque', 'Type Moto', 'Chassis', 'Date', 
            'Propriétaire', 'Contact Propriétaire',
            'Client', 'Contact Client', 'No Fiche', 'Couleur', 'Cellulaire'
        ])
        
        for obj in queryset:
            writer.writerow([
                obj.immatriculation,
                obj.marque.nom if obj.marque else '',
                obj.type_moto.nom if obj.type_moto else '',
                obj.chassis,
                obj.date,
                obj.proprietaire.nom if obj.proprietaire else '',
                obj.proprietaire.contact if obj.proprietaire else '',
                obj.client.nom if obj.client else '',
                obj.client.contact if obj.client else '',
                obj.no_fiche,
                obj.couleur,
                obj.cellulaire
            ])
        
        return response

    def export_all_pdf(self, request, queryset):
        """Export PDF complet avec toutes les données détaillées pour les motos"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from django.http import HttpResponse

        buffer = BytesIO()
        # Marges minimales pour maximiser l'espace d'affichage
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              leftMargin=0.05*inch, rightMargin=0.05*inch, 
                              topMargin=0.1*inch, bottomMargin=0.1*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=12,
            spaceAfter=10,
            alignment=1
        )
        title = Paragraph("Export Complet - Motos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 10))

        # En-têtes détaillés SANS Statut et SANS adresses
        headers = [
            'Immatriculation', 'Marque', 'Type Moto', 'Chassis', 'Date', 
            'Propriétaire', 'Contact Propriétaire',
            'Client', 'Contact Client', 'No Fiche', 'Couleur', 'Cellulaire'
        ]
        
        cell_style = ParagraphStyle('cell', fontSize=6, leading=7, alignment=1)
        data = [headers]
        for obj in queryset:
            data.append([
                Paragraph(obj.immatriculation, cell_style),
                Paragraph(truncate_text(obj.marque.nom if obj.marque else '-', 10), cell_style),
                Paragraph(truncate_text(obj.type_moto.nom if obj.type_moto else '-', 12), cell_style),
                Paragraph(truncate_text(obj.chassis, 15), cell_style),
                Paragraph(str(obj.date)[:10] if obj.date else '-', cell_style),
                Paragraph(truncate_text(obj.proprietaire.nom if obj.proprietaire else '-', 15), cell_style),
                Paragraph(truncate_text(obj.proprietaire.contact if obj.proprietaire else '-', 15), cell_style),
                Paragraph(truncate_text(obj.client.nom if obj.client else '-', 15), cell_style),
                Paragraph(truncate_text(obj.client.contact if obj.client else '-', 15), cell_style),
                Paragraph(truncate_text(obj.no_fiche, 10), cell_style),
                Paragraph(truncate_text(obj.couleur, 10), cell_style),
                Paragraph(truncate_text(obj.cellulaire, 15), cell_style)
            ])
        
        # Calculer automatiquement les largeurs de colonnes pour s'adapter à la page
        available_width = landscape(A4)[0] - 0.1*inch  # Largeur disponible moins les marges
        num_columns = len(headers)
        base_width = available_width / num_columns
        
        # Ajuster certaines colonnes selon leur contenu
        col_widths = []
        for i, header in enumerate(headers):
            if 'Immatriculation' in header:
                col_widths.append(base_width * 1.5)  # Plus large pour l'immatriculation
            elif 'Chassis' in header:
                col_widths.append(base_width * 1.3)
            elif 'Date' in header:
                col_widths.append(base_width * 0.8)
            else:
                col_widths.append(base_width)
        
        table = Table(data, colWidths=col_widths)
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(table_style)
        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="motos_complet_export.pdf"'
        response.write(pdf)
        return response

    def export_all_excel(self, request, queryset):
        """Export Complet Excel au format personnalisé et coloré pour les motos"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Motos"
        # En-têtes personnalisés
        headers = ['NOM DU PROPRIÉTAIRE', 'CONTACT PROPRIÉTAIRE', 'MARQUE', 'TYPE', 'N° CHASSIS', 'N° IMMAT.', 'DATES', 'NO FICHE', 'CLIENT', 'CONTACT CLIENT', 'COULEURS']
        ws.append(headers)
        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Données
        immat_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Jaune clair
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.proprietaire.nom if obj.proprietaire and obj.proprietaire.nom else '')
            ws.cell(row=row_idx, column=2, value=obj.proprietaire.contact if obj.proprietaire and obj.proprietaire.contact else '')
            ws.cell(row=row_idx, column=3, value=obj.marque.nom if obj.marque and obj.marque.nom else '')
            ws.cell(row=row_idx, column=4, value=obj.type_moto.nom if obj.type_moto and obj.type_moto.nom else '')
            ws.cell(row=row_idx, column=5, value=obj.chassis if obj.chassis else '')
            ws.cell(row=row_idx, column=6, value=obj.immatriculation if obj.immatriculation else '')
            ws.cell(row=row_idx, column=7, value=obj.date.strftime('%d/%m/%Y') if obj.date else '')
            ws.cell(row=row_idx, column=8, value=obj.no_fiche if obj.no_fiche else '')
            ws.cell(row=row_idx, column=9, value=obj.client.nom if obj.client and obj.client.nom else '')
            ws.cell(row=row_idx, column=10, value=obj.client.contact if obj.client and obj.client.contact else '')
            ws.cell(row=row_idx, column=11, value=obj.couleurs if obj.couleurs else '')
            # Appliquer le fond jaune à la colonne N° IMMAT.
            ws.cell(row=row_idx, column=6).fill = immat_fill
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="motos_complet_export.xlsx"'
        return response

# Personnalisation de l'interface d'administration
admin.site.site_header = mark_safe('<img src="/static/images/logo.png" alt="Logo" style="height:48px;vertical-align:middle;margin-right:12px;"> Gestion Véhicules - Administration Professionnelle')
admin.site.site_title = "Gestion Véhicules"
admin.site.index_title = "📊 Tableau de bord - Gestion des Véhicules"
admin.site.site_url = None

# Ajouter les statistiques au contexte de l'admin
def get_admin_stats():
    """Récupérer les statistiques détaillées pour l'admin"""
    try:
        from django.db import connection
        
        # Vérifier si les tables existent
        with connection.cursor() as cursor:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='vehicules_vehicule';")
            table_exists = cursor.fetchone()
            
        if not table_exists:
            # Si les tables n'existent pas, retourner des statistiques vides
            return {
                'vehicules_count': 0,
                'motos_count': 0,
                'clients_count': 0,
                'proprietaires_count': 0,
                'marques_count': 0,
                'types_vehicule_count': 0,
                'types_moto_count': 0,
                'vehicules_ce_mois': 0,
                'motos_ce_mois': 0,
                'nouveaux_clients_ce_mois': 0,
                'top_marques': [],
                'activite_recente': {
                    'vehicules_recent': 0,
                    'motos_recent': 0,
                }
            }
        
        # Si les tables existent, calculer les statistiques normalement
        today = timezone.now().date()
        this_month = today.replace(day=1)
        this_year = today.replace(month=1, day=1)
        
        return {
            'vehicules_count': Vehicule.objects.count(),
            'motos_count': Moto.objects.count(),
            'clients_count': Client.objects.count(),
            'proprietaires_count': Proprietaire.objects.count(),
            'marques_count': Marque.objects.count(),
            'types_vehicule_count': TypeVehicule.objects.count(),
            'types_moto_count': TypeMoto.objects.count(),
            
            # Statistiques temporelles
            'vehicules_ce_mois': Vehicule.objects.filter(date__gte=this_month).count(),
            'motos_ce_mois': Moto.objects.filter(date__gte=this_month).count(),
            'nouveaux_clients_ce_mois': Client.objects.count(),  # Simplifié car pas de date_creation
            
            # Top marques
            'top_marques': Marque.objects.annotate(
                total_vehicules=Count('vehicule') + Count('moto')
            ).order_by('-total_vehicules')[:5],
            
            # Activité récente
            'activite_recente': {
                'vehicules_recent': Vehicule.objects.filter(date__gte=today - timedelta(days=7)).count(),
                'motos_recent': Moto.objects.filter(date__gte=today - timedelta(days=7)).count(),
            }
        }
    except Exception as e:
        # En cas d'erreur, retourner des statistiques vides
        print(f"Error in get_admin_stats: {e}")
        return {
            'vehicules_count': 0,
            'motos_count': 0,
            'clients_count': 0,
            'proprietaires_count': 0,
            'marques_count': 0,
            'types_vehicule_count': 0,
            'types_moto_count': 0,
            'vehicules_ce_mois': 0,
            'motos_ce_mois': 0,
            'nouveaux_clients_ce_mois': 0,
            'top_marques': [],
            'activite_recente': {
                'vehicules_recent': 0,
                'motos_recent': 0,
            }
        }

# Surcharger le contexte de l'admin
from django.contrib.admin.sites import AdminSite
original_each_context = AdminSite.each_context

def custom_each_context(self, request):
    context = original_each_context(self, request)
    context.update(get_admin_stats())
    return context

AdminSite.each_context = custom_each_context 